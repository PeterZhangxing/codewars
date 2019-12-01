import smtplib,os,datetime,time
import pymysql as pms
import openpyxl,uuid


def conn_database(*args,**kwargs):
    '''
    连接指定的数据库
    :param args:
    :param kwargs:
    :return: 成功返回数据库连接，失败返回错误信息
    '''
    try:
        res = pms.connect(**kwargs)
    except Exception as e:
        res = str(e)
    return res


def get_datas(sql):
    '''
    连接数据库，连接成功则获取数据库查询结果中每个字段的名字和每行数据的值，不成功则退出整个程序
    :param sql:
    :return:
    '''
    # conn = pms.connect(host='192.168.31.235',user='ormuser',passwd='redhat',database='ormtest',port=3306,charset='utf8')
    conn = conn_database(
        host='192.168.31.235',
        user='ormuser',
        passwd='redhat',
        database='ormtest',
        port=3306,
        charset='utf8'
    )
    if isinstance(conn,str):
        exit(conn)
    else:
        print("connected with the database")
        cur = conn.cursor()
        cur.execute(sql)
        datas = cur.fetchall()
        '''
        从数据库中获取的数据格式如下(二元元组)：
        ((1, 'bobo', 27, 0, 1, 1, 'py1', 12000), (2, 'jiji', 34, 0, 2, 2, 'py2', 12000),...)
        '''
        fields = cur.description
        '''
        从数据库中获取字段的属性，格式如下所示(二元元组)：
        (('id', 3, None, 11, 11, 0, False), ('name', 253, None, 32, 32, 0, False),...)
        '''
        cur.close()
        res_dic = {'datas':datas,'fields':fields}
        return res_dic


# def get_fields(sql):
#     conn = pms.connect(host='192.168.31.235',user='ormuser',passwd='redhat',database='ormtest',port=3306,charset='utf8')
#     cur = conn.cursor()
#     cur.execute(sql)
#     fields = cur.description
#     '''
#     从数据库中获取字段的属性，格式如下所示(二元元组)：
#     (('id', 3, None, 11, 11, 0, False), ('name', 253, None, 32, 32, 0, False),...)
#     '''
#     cur.close()
#     return fields


def get_excel(data,field,file):
    '''
    根据从数据库中获取的字段名和字段值，生产excel表格
    :param data:
    :param field:
    :param file:
    :return:
    '''
    sheet_id = str(uuid.uuid4())[:5]
    new = openpyxl.Workbook() # 生成一个在内存中的excel空表格
    sheet = new.active # 生成表格中的一个新表单
    sheet.title = "%s_sheet"%(sheet_id) # 为新表单取个名字

    # 生成表格中每列的名称
    for col in range(len(field)):
        _ = sheet.cell(row=1,column=col+1,value=u'%s'%(field[col][0]))

    # 生成表格中的每行数据
    for row in range(len(data)):
        for col in range(len(field)):
            _ = sheet.cell(row=row+2,column=col+1,value=u'%s'%(data[row][col]))

    # 将生成的表格保存到指定的文件中
    my_work_book = new.save(file)

    return my_work_book


def getdate():
    '''
    获取昨天的日期
    :return:
    '''
    today = datetime.date.today() # 获取今天的时间戳
    oneday = datetime.timedelta(days=1) # 获取一天时间的时间戳长度
    yesterday = today - oneday # 得到昨天的时间戳表示
    yesstr = yesterday.strftime('%Y-%m-%d') # 获取字符时间输出
    return yesstr


def cexcel_as_required(sql):
    '''
    调用函数，获取数据库数据，并生成excel表格
    :param sql:
    :return:
    '''
    # 输入自定义生成的excel文件的名称
    sd_file = input('Please input the name of excel you want to create here: ')
    file_name = '%s_%s.xlsx' % (sd_file, getdate())

    # 生成存放excel文件的绝对路径
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))
    FILE_PATH = os.path.join(BASE_DIR,'statics')
    file = os.path.join(FILE_PATH,file_name)

    try:
        # 获取数据库中的数据
        res = get_datas(sql)
        data = res.get('datas')
        field = res.get('fields')

        # 生成表格
        get_excel(data,field,file)

    except Exception as e:
        exit(str(e))


if __name__ == "__main__":

    # 定义用于查询的sql语句
    sql = 'select * from regulation_students left join regulation_classes on regulation_students.stu_cla_id=regulation_classes.id;'
    cexcel_as_required(sql)