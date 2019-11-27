import openpyxl
from openpyxl.styles import Font,PatternFill,colors,Alignment


# # 读取excel表格的内容，并保存到对象中
# my_wb = openpyxl.load_workbook('test.xlsx')
#
# # 获取当前excel表格的所有sheet的名字
# all_sheets = my_wb.get_sheet_names()
# print(all_sheets) # ['students', 'classes', 'Sheet3']
#
# # 获取当前活动的sheet，就是可以被操作的sheet
# my_active_sheet = my_wb.get_active_sheet()
# print(my_active_sheet) # <Worksheet "classes">
#
# # 通过sheet的名字，获取到sheet对象
# my_sheet = my_wb.get_sheet_by_name('students')
# print(my_sheet) # <Worksheet "students">
#
# print(my_sheet.max_row) # 获取表格的最大行数
# print(my_sheet.max_column) # 获取表格的最大列数
#
# # 遍历表格中所有的指定的单元格
# for i in range(2,my_sheet.max_row+1):
#     for j in range(2,my_sheet.max_column+1):
#         col_name = my_sheet.cell(row=1,column=j).value
#         cell_content = my_sheet.cell(row=i,column=j).value
#         print(cell_content,end=' ')
#
#         if cell_content is None and col_name == '年龄':
#             # 给某个单元格赋值
#             my_sheet.cell(row=i, column=j).value = 100
#
#         username = my_sheet.cell(row=i,column=j-1).value
#         if username=='zx' and col_name == '年龄':
#             # 给某个单元格赋值
#             my_sheet.cell(row=i, column=j).value = 100
#     print()
#
# # 保存修改后的文件到指定的路径中
# my_wb.save('backup.xlsx')

# # 创建一个新的excel表格文件,，自带一个sheet
# new_wb = openpyxl.Workbook()
# # print(new_wb.get_sheet_names()) # ['Sheet']
#
# # 创建一个空白的sheet
# sheet1 = new_wb.create_sheet()
# # print(new_wb.get_sheet_names()) # ['Sheet', 'Sheet1']
#
# # 创建一个自定义title的sheet
# sheet2 = new_wb.create_sheet('railwaystation')
# # print(new_wb.get_sheet_names()) # ['Sheet', 'Sheet1', 'railwaystation']
#
# old_new_sheet_dict = {}
# title_li = ['hospitals','trainstations','airports']
# for i in range(len(new_wb.get_sheet_names())):
#     old_new_sheet_dict[new_wb.get_sheet_names()[i]] =  title_li[i]
#
# for k,v in old_new_sheet_dict.items():
#     sheet = new_wb.get_sheet_by_name(k)
#     sheet.title = v # 修改sheet的值
#
# new_wb.save('change_sheetname.xlsx')

# 读取excel表格的内容，并保存到对象中
# my_wb = openpyxl.load_workbook('test.xlsx')
#
# sheet_name_li = my_wb.get_sheet_names()
#
# for name in sheet_name_li:
#     sheet_obj = my_wb.get_sheet_by_name(name)
#     for i in range(2,sheet_obj.max_row+1):
#         for j in range(1,sheet_obj.max_column+1):
#             print(sheet_obj.cell(row=i,column=j).value,end=' ')
#         print()
#     print("*"*100)


new_wb = openpyxl.Workbook()

# 设置字体对象的属性
myfont = Font(size=24,bold=True,italic=True,color=colors.GREEN)
# 设置填充对象的属性
myfill = PatternFill(fill_type="solid",fgColor= colors.RED)
# 设置位置对象的属性
myalignment = Alignment(horizontal='center',vertical='center')

# mysheet = new_wb.get_active_sheet()
mysheet = new_wb.active

# 设置行高
mysheet.row_dimensions[1].height = 70
# 设置列宽
mysheet.column_dimensions['A'].width = 100

# 设置某单元格的字体样式
mysheet['A1'].font = myfont
# 设置某单元格的内容
mysheet['A1'] = 'fuck you'

# 设置某单元格的字体样式
mysheet.cell(row=2,column=1).font = myfont
# 设置单元格的填充色
mysheet.cell(row=2,column=1).fill = myfill
# 设置单元格中文字的位置
mysheet.cell(row=2,column=1).alignment = myalignment
# 设置某单元格的内容
mysheet.cell(row=2,column=1).value = 'your mom'

# 合并单元格
mysheet.merge_cells(start_row=3,start_column=1,end_row=5,end_column=3)
mysheet.merge_cells('A8:C12')

new_wb.save('testfont.xlsx')